from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#load existing spreadsheet#
wb = load_workbook('C:\\Users\\johnb\\Desktop\\hello.xlsx')

#create an active worksheet
ws = wb.active

#print something from the spreadsheet
print(f'{ws["A2"].value}: {ws["B2"].value}')

